from openpyxl import Workbook as p
import openpyxl as P
from openpyxl.workbook import Workbook
# from openpyxl.workbook import load_workbook
from os import path

"""
wb = Workbook()
ws = wb.active
wb.title = "Orgo"

ws.title = "Chapter 1"

ws1 = wb.create_sheet("Chapter 2")
ws2 = wb.create_sheet("Chapter 3")
ws3 = wb.create_sheet("Chapter 4")

d = ws.cell(4, 1, "Ikhla3")

for row in ws.values:
    for value in row:
        print(value)

for x in range(1, 4):
    for y in range(1, 4):
        ws.cell(x, y)


def load_workbook(wb_path):
    if path.exists(wb_path):
        return P.load_workbook(wb_path)
    return P.Workbook()


wb.save("scratch_file.xlsx")
wb2 = load_workbook("Microbiology.xlsx")

for row in ws.iter_rows(min_row=1, max_col=3, max_row=4, values_only=True):
    print(row)


for sheet in wb:
    print(sheet.title)
"""


def new_workbook(namewb: str, namews: str):
    '''Makes a new workbook for a new subject you would like to be tested on. Allows you to name the workbook and
    your first sheet.'''
    wb = Workbook()
    ws = wb.active
    ws.title = namews
    wb.save(namewb + ".xlsx")


"""
new_workbook("Microbiology", "Chapter1")
"""  # This function has been tested and works because it created mirobiology


def add_worksheet(namewb: str, namews: str):
    """Adds a new worksheet to an existing workbook. You do not need to add .xslx at the end of your workbook name :)))"""
    # My initial struggle is making it understand WHICH workbook to take if it exists within the computer
    actnamewb = namewb + ".xlsx"
    if path.exists(actnamewb):
        wb = P.load_workbook(actnamewb)
    wb.create_sheet(namews)
    wb.save(actnamewb)


"""
add_worksheet("Microbiology", "Chapter2" )
"""  # This function works becasuse it created a sheet called chapter2 in microbiology


def change_worksheet_name(namewb: str, namews: str, newwsname: str):
    ...

def copy_workbook(namewb: str):
    if path.exists(namewb + ".xlsx"):
        copywb = P.load_workbook(namewb + ".xlsx")
    namewb.save(copywb + ".xlsx")




def change_workbook_name(namewb: str, newwbname: str):
    """Replaces the workbook name with a new name. No need to add .xlsx at the end"""
    if path.exists(namewb + ".xlsx"):
        namewb = P.load_workbook(namewb + ".xlsx")
    namewb.title = newwbname
    wb.save(newwbname + ".xlsx")




#change_workbook_name("Microbiology", "MicroBio") ##### Has a bug where it simply creates a new workbook called MicroBio with the same sheets as Microbiology

#copy_workbook("MicroBio") #Gives an error I am not sure how to continue

def new_question(question: str):
    ...


def new_ans():
    ...
